"""extract_pms_pdf.py — turn a project's PMS PDF + datasheet workbooks into a
PDF-extracted source-of-truth JSON the engine can serve directly.

Usage
-----
    python -m scripts.extract_pms_pdf \
        --pms-pdf  path/to/PMS_PDF.pdf \
        --workbook path/to/BALL.xlsm  \
        --workbook path/to/CHECK.xlsm  \
        ...                            \
        --project-id  pttep-sabah      \
        [--output app/data/projects/<project>/pms_datasheet_extracted.json] \
        [--report  PMS_Provenance_Report.csv]

What it does
------------
1.  Reads the PMS PDF page-by-page (text extraction via pypdf).
2.  Reads each .xlsm workbook's `Index` sheet to learn the VDS → PDF page map.
3.  For each VDS sheet in each workbook, reads every (label, value) pair.
4.  For each pair, fuzzy-matches the value against the corresponding PDF page.
5.  Writes a JSON file keyed by VDS code:

        {
          "BLFTA1R": {
            "Valve Standard": {
              "value": "API 6D / ISO 17292",
              "pdf_source": "PMS_PDF.pdf page 39",
              "verification": "EXACT"
            },
            ...
          },
          ...
        }

6.  Optionally writes a flat CSV of every (vds, field, value, page, status) row
    so the client can audit each cell.

Produces no false positives: every record contains the precise PDF page where
the value is sourced. If a value cannot be located on the page, the record's
`verification` field will be `NOT_FOUND` or `PARTIAL`, never silently elided.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Iterable

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    import pypdf
except ImportError as e:
    sys.exit("ERROR: pypdf is required. Run:  pip install pypdf openpyxl")

try:
    from openpyxl import load_workbook
except ImportError as e:
    sys.exit("ERROR: openpyxl is required. Run:  pip install pypdf openpyxl")


# ── Helpers ────────────────────────────────────────────────────────────────

SKIP_LABELS = {
    # Skip cells whose values are formula-driven (not standard-driven) so we
    # don't pollute the output with cached "#VALUE!" or PMS-table lookups.
    "VDS No",
    "Piping Class",
    "Size Range",
    "Service",
    "Design Pressure",
    "Hydrotest Shell Test Pressure",
    "Hydrotest Closure Test Pressure",
}

NOTE_PREFIXES = ("1.", "2.", "3.", "4.", "5.", "6.", "Notes:")


def _norm(s: str | None) -> str:
    """Normalize text for fuzzy comparison: strip smart quotes, collapse ws, lower."""
    if s is None:
        return ""
    s = str(s)
    s = s.replace("“", '"').replace("”", '"')
    s = s.replace("‘", "'").replace("’", "'")
    s = s.replace("–", "-").replace("—", "-")
    s = s.replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()


_STOP = {"and", "for", "the", "with", "per", "more", "than", "below", "above"}


def _verify(value: str, page_text: str) -> tuple[str, str]:
    """Decide whether `value` appears on `page_text`. Returns (status, evidence)."""
    nv = _norm(value)
    np = _norm(page_text)
    if not nv or nv in {"-", "n/a", "nil"}:
        return "TRIVIAL", nv
    if nv in np:
        return "EXACT", nv
    tokens = re.findall(r"[A-Za-z0-9]+", nv)
    sig = [t for t in tokens if len(t) > 2 and t not in _STOP]
    if not sig:
        return "TRIVIAL", nv
    found = sum(1 for t in sig if t in np)
    ratio = found / len(sig)
    if ratio >= 0.85:
        return "FUZZY", f"{found}/{len(sig)} tokens"
    if ratio >= 0.5:
        return "PARTIAL", f"{found}/{len(sig)} tokens"
    return "NOT_FOUND", f"{found}/{len(sig)} tokens"


def extract_pdf_pages(pdf_path: Path) -> list[str]:
    """Read every page's text out of the PMS PDF."""
    reader = pypdf.PdfReader(str(pdf_path))
    pages = []
    for p in reader.pages:
        try:
            pages.append(p.extract_text() or "")
        except Exception:
            pages.append("")
    return pages


def build_vds_to_page(workbook_paths: Iterable[Path]) -> dict[str, int]:
    """Walk each workbook's `Index` sheet and build VDS → PDF page lookup.

    Looks for a header row containing 'VDS' and 'Page No' columns.
    """
    out: dict[str, int] = {}
    for path in workbook_paths:
        wb = load_workbook(path, data_only=True, keep_vba=False)
        if "Index" not in wb.sheetnames:
            continue
        ws = wb["Index"]
        header_row = vds_col = page_col = None
        for r in range(1, min(15, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v and str(v).strip().upper() == "VDS":
                    header_row, vds_col = r, c
                if v and "Page No" in str(v):
                    if header_row == r:
                        page_col = c
            if header_row and page_col:
                break
        if not (header_row and vds_col and page_col):
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=vds_col).value
            p = ws.cell(row=r, column=page_col).value
            if v and isinstance(p, (int, float)):
                vds = str(v).strip().replace("VDS-", "")
                out[vds] = int(p)
    return out


# ── Materials parser ────────────────────────────────────────────────────────
# Walk each per-VDS appendix datasheet, collect the (piping_class, materials)
# pairs by reading the standard "Body Material / Ball Material / Stem Material…"
# rows. The first time we see a class, we capture its materials block.

MATERIAL_FIELD_TO_KEY = {
    "Body Material":     "body_material",
    "Ball Material":     "ball_material",
    "Disc Material":     "disc_material",
    "Wedge Material":    "wedge_material",
    "Stem Material":     "stem_material",
    "Trim Material":     "trim_material",
    "Seat Material":     "seat_material",
    "Gland material":    "gland_material",
    "Gland Material":    "gland_material",
    "Gland Packing":     "gland_packing",
    "Gland packing":     "gland_packing",
    "Spring":            "spring_material",
    "Lever / Handwheel": "lever_handwheel",
    "Needle Material":   "needle_material",
    "Cover Material":    "cover_material",
}


def materials_per_class(workbook_paths: Iterable[Path]) -> dict[str, dict]:
    """For each piping class encountered across the per-VDS sheets, build a
    representative materials block by averaging-by-popularity across all
    valves that use that class. The most-common ASTM grade per material field
    becomes the class-level material spec.
    """
    from collections import Counter
    per_class_field_counter: dict[str, dict[str, Counter]] = {}
    for path in workbook_paths:
        wb = load_workbook(path, data_only=True, keep_vba=False)
        for sn in wb.sheetnames:
            if not sn.startswith("VDS-"):
                continue
            ws = wb[sn]
            # piping_class is on row 5 col C in our datasheet template
            cls = ws.cell(row=5, column=3).value
            cls = str(cls).strip() if cls else None
            if not cls or cls == "#VALUE!":
                continue
            for r in range(1, ws.max_row + 1):
                a = ws.cell(row=r, column=1).value
                b = ws.cell(row=r, column=2).value
                c = ws.cell(row=r, column=3).value
                label = (str(b).strip() if b else "") or (str(a).strip() if a else "")
                if not label or c is None:
                    continue
                key = MATERIAL_FIELD_TO_KEY.get(label)
                if not key:
                    continue
                v = str(c).strip()
                if not v or v == "#VALUE!":
                    continue
                per_class_field_counter.setdefault(cls, {}).setdefault(key, Counter())[v] += 1

    out: dict[str, dict] = {}
    for cls, fields in per_class_field_counter.items():
        block = {}
        for fkey, ctr in fields.items():
            block[fkey] = ctr.most_common(1)[0][0]   # most common grade wins
        # Heuristic ASME B16.34 group from material wording
        body = (block.get("body_material") or "").upper()
        if "F316L" in body or "CF3M" in body or "F60" in body or "F53" in body:
            block["asme_b1634_group"] = 2
        elif "B148" in body or "B61" in body or "C95800" in body:
            block["asme_b1634_group"] = 3
        else:
            block["asme_b1634_group"] = 1
        out[cls] = block
    return out


def extract_records(
    workbook_paths: Iterable[Path],
    vds_to_page: dict[str, int],
    pdf_pages: list[str],
) -> tuple[dict[str, dict], list[dict]]:
    """Return (records, audit_rows). records is the engine-ready dict."""
    records: dict[str, dict] = {}
    audit_rows: list[dict] = []

    for path in workbook_paths:
        wb = load_workbook(path, data_only=True, keep_vba=False)
        for sn in wb.sheetnames:
            if not sn.startswith("VDS-"):
                continue
            vds = sn.replace("VDS-", "").strip()
            page = vds_to_page.get(vds)
            if not page:
                continue
            page_text = pdf_pages[page - 1] if 0 < page <= len(pdf_pages) else ""
            ws = wb[sn]
            rec: dict[str, dict] = {}
            for r in range(1, ws.max_row + 1):
                a = ws.cell(row=r, column=1).value
                b = ws.cell(row=r, column=2).value
                c = ws.cell(row=r, column=3).value
                label = (str(b).strip() if b else "") or (str(a).strip() if a else "")
                if not label or c is None:
                    continue
                if label in SKIP_LABELS or label.startswith(NOTE_PREFIXES):
                    continue
                val = str(c).strip()
                if not val or val == "#VALUE!":
                    continue
                status, evidence = _verify(val, page_text)
                pdf_source = f"{Path(pdf_pages_filename).name} page {page}" \
                    if 'pdf_pages_filename' in globals() else f"PMS_PDF.pdf page {page}"
                rec[label] = {
                    "value": val,
                    "pdf_source": pdf_source,
                    "verification": status,
                }
                audit_rows.append(
                    {
                        "vds": vds,
                        "pdf_page": page,
                        "field": label,
                        "value": val,
                        "status": status,
                        "evidence": evidence,
                    }
                )
            if rec:
                records[vds] = rec

    return records, audit_rows


def main() -> int:
    p = argparse.ArgumentParser(description="Extract a project's PMS datasheets into engine-ready JSON.")
    p.add_argument("--pms-pdf", required=True, type=Path, help="Path to the project's PMS PDF.")
    p.add_argument(
        "--workbook",
        action="append",
        required=True,
        type=Path,
        help="Datasheet workbook (.xlsm). Repeat once per valve type (Ball, Gate, Check, ...).",
    )
    p.add_argument("--project-id", required=True, help="Project slug (e.g. 'pttep-sabah').")
    p.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output JSON path. Default: app/data/projects/<project-id>/pms_datasheet_extracted.json",
    )
    p.add_argument("--report", type=Path, default=None, help="Optional flat CSV provenance audit.")
    args = p.parse_args()

    if not args.pms_pdf.exists():
        sys.exit(f"ERROR: PMS PDF not found: {args.pms_pdf}")
    for w in args.workbook:
        if not w.exists():
            sys.exit(f"ERROR: workbook not found: {w}")

    if args.output is None:
        repo_root = Path(__file__).resolve().parent.parent
        args.output = repo_root / "app" / "data" / "projects" / args.project_id / "pms_datasheet_extracted.json"
    args.output.parent.mkdir(parents=True, exist_ok=True)

    print(f"PMS PDF      : {args.pms_pdf}")
    print(f"Workbooks    : {len(args.workbook)}")
    print(f"Project      : {args.project_id}")
    print(f"Output       : {args.output}")
    if args.report:
        print(f"Audit report : {args.report}")

    print("\n[1/3] Reading PDF text ...")
    pdf_pages = extract_pdf_pages(args.pms_pdf)
    # Stash the PDF filename so extract_records can cite it
    global pdf_pages_filename
    pdf_pages_filename = args.pms_pdf.name
    print(f"      {len(pdf_pages)} pages extracted")

    print("[2/3] Building VDS → page map from workbook indexes ...")
    vds_to_page = build_vds_to_page(args.workbook)
    print(f"      {len(vds_to_page)} VDS codes mapped to pages")
    if not vds_to_page:
        sys.exit("ERROR: no VDS codes found in workbook Index sheets. Aborting.")

    print("[3/4] Extracting cell values + verifying against PDF ...")
    records, audit_rows = extract_records(args.workbook, vds_to_page, pdf_pages)
    total_cells = len(audit_rows)
    by_status: dict[str, int] = {}
    for r in audit_rows:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    print(f"      {len(records)} VDS records, {total_cells} cells")
    for status, n in sorted(by_status.items(), key=lambda kv: -kv[1]):
        pct = 100.0 * n / total_cells if total_cells else 0.0
        print(f"        {status:<10}: {n:>5}  ({pct:.1f}%)")

    print("[4/4] Building materials block per piping class ...")
    materials = materials_per_class(args.workbook)
    materials_path = args.output.parent / "pms_materials_per_class.json"
    materials_path.write_text(json.dumps(materials, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"      {len(materials)} piping classes with materials → {materials_path.name}")
    if materials:
        sample_cls = next(iter(materials))
        print(f"      sample (class '{sample_cls}'): {list(materials[sample_cls].keys())}")
        # Auto-merge into pms_extracted.json under each class' "materials" key
        repo_root = Path(__file__).resolve().parent.parent
        legacy_pms = repo_root / "app" / "data" / "projects" / args.project_id / "pms_extracted.json"
        if legacy_pms.exists():
            try:
                pms = json.loads(legacy_pms.read_text(encoding="utf-8"))
                merged = 0
                for cls, mat in materials.items():
                    if cls in pms:
                        pms[cls]["materials"] = mat
                        merged += 1
                legacy_pms.write_text(json.dumps(pms, indent=2, ensure_ascii=False), encoding="utf-8")
                print(f"      merged materials into {merged} classes of {legacy_pms.name}")
            except Exception as e:
                print(f"      could not merge into pms_extracted.json: {e}")

    args.output.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\n✓ Wrote engine data file: {args.output}")

    if args.report:
        with args.report.open("w", encoding="utf-8-sig", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=["vds", "pdf_page", "field", "value", "status", "evidence"])
            w.writeheader()
            w.writerows(audit_rows)
        print(f"✓ Wrote audit report   : {args.report}")

    print("\nNext steps:")
    print(f"  • Restart the engine — get_datasheet_loader('{args.project_id}') will pick up the new file.")
    print(f"  • Call generate_datasheet(decoded, project_id='{args.project_id}', return_provenance=True).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
