"""XLSX parser for project PMS sheets.

Section-aware parser. Each worksheet = one piping class. Sections are
detected by single-cell headers in column A (e.g. "Pipe Data", "Valves",
"Flange"). Within a section, a row whose column-B label is "Size (in)"
becomes the column→NPS map for subsequent rows.

Designed to handle the PMS_B1N_300.xlsx layout but tolerant of variations:
unknown rows land in `extra` rather than being dropped.
"""
from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from .schema import (
    AttributeValue,
    PipingClass,
    ProjectMetadata,
    ProjectPMS,
    PTRating,
    PipeScheduleRow,
    ValveAssignment,
)

# ---- value helpers ----------------------------------------------------------

_NUM_RE = re.compile(r"-?\d+(?:\.\d+)?")


def _to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = _NUM_RE.search(str(v).replace(",", ""))
    return float(m.group()) if m else None


def _tokens(v: Any) -> List[str]:
    if v is None:
        return []
    s = str(v).lower()
    return [t for t in re.split(r"[,/;\s\-]+", s) if t and len(t) > 1]


def _attr(v: Any, unit: Optional[str] = None) -> AttributeValue:
    return AttributeValue(
        raw=v if not isinstance(v, str) else v.strip(),
        numeric=_to_float(v),
        unit=unit,
        tokens=_tokens(v),
    )


def _is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


# ---- header normalization ---------------------------------------------------

HEADER_KEY_MAP = {
    "piping class": "spec_code",
    "rating": "pressure_rating",
    "material": "material_description",
    "corrosion allowance": "corrosion_allowance",
    "mill tolerance": "mill_tolerance",
    "design code": "design_code",
    "service": "service",
    "branch chart": "branch_chart",
}


def _normalize_key(label: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "_", label.strip().lower()).strip("_")
    return HEADER_KEY_MAP.get(label.strip().lower(), s)


# ---- section detection ------------------------------------------------------

SECTION_PATTERNS = [
    ("header",            re.compile(r"^piping\s*material\s*specification", re.I)),
    ("pt_ratings",        re.compile(r"pressure[\s\-]*temperature\s*rating", re.I)),
    ("pipe_data",         re.compile(r"^pipe\s*data", re.I)),
    ("fittings_bw",       re.compile(r"fittings.*butt\s*weld", re.I)),
    ("fittings_extra",    re.compile(r"^extra\s*fittings", re.I)),
    ("flange",            re.compile(r"^flange", re.I)),
    ("blinds",            re.compile(r"spectacle.*blind|spacer\s*blind", re.I)),
    ("bolts",             re.compile(r"bolts?.*nuts?.*gasket", re.I)),
    ("valves",            re.compile(r"^valves\b", re.I)),
    ("notes",             re.compile(r"^notes\b", re.I)),
]


def _detect_section(cell_a: Any) -> Optional[str]:
    if _is_blank(cell_a):
        return None
    s = str(cell_a)
    for name, pat in SECTION_PATTERNS:
        if pat.search(s):
            return name
    return None


# ---- per-sheet parser -------------------------------------------------------

VALVE_TYPE_MAP = {
    "ball": "BALL", "gate": "GATE", "globe": "GLOBE",
    "check": "CHECK", "butterfly": "BUTTERFLY",
    "needle": "NEEDLE", "dbb": "DBB", "plug": "PLUG",
}


def _row_values(ws, r: int) -> List[Any]:
    return [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]


def parse_sheet(ws) -> PipingClass:
    pc = PipingClass(spec_code=ws.title.strip())
    section: Optional[str] = "header"
    size_cols: Dict[int, float] = {}      # excel column index (1-based) -> NPS
    pt_temp_row: Optional[Dict[int, float]] = None
    notes: List[str] = []

    for r in range(1, ws.max_row + 1):
        row = _row_values(ws, r)
        if all(_is_blank(v) for v in row):
            continue
        cell_a = row[0]
        cell_b = row[1] if len(row) > 1 else None

        new_sec = _detect_section(cell_a)
        if new_sec:
            section = new_sec
            size_cols = {}
            pt_temp_row = None
            # capture spec_code label too
            if section == "header" and not _is_blank(cell_a):
                pass
            continue

        # Within "header" section: column A is label, column C is value
        if section == "header":
            label = cell_a
            value = row[2] if len(row) > 2 else cell_b
            if _is_blank(label) or _is_blank(value):
                continue
            key = _normalize_key(str(label))
            if key == "spec_code" and isinstance(value, str):
                pc.spec_code = value.strip()
                continue
            pc.attributes[key] = _attr(value)
            continue

        # Detect a "Size (in)" row -> build size column map
        label_b = str(cell_b).strip().lower() if isinstance(cell_b, str) else ""
        if "size" in label_b and "in" in label_b:
            size_cols = {}
            for ci in range(3, len(row) + 1):
                v = row[ci - 1]
                f = _to_float(v)
                if f is not None:
                    size_cols[ci] = f
            continue

        # PT ratings: capture Temp row and Press row
        if section == "pt_ratings":
            if isinstance(cell_b, str) and "temp" in cell_b.lower() and "press" not in cell_b.lower():
                pt_temp_row = {}
                for ci in range(3, len(row) + 1):
                    f = _to_float(row[ci - 1])
                    if f is not None:
                        pt_temp_row[ci] = f
                continue
            if isinstance(cell_b, str) and "press" in cell_b.lower():
                if pt_temp_row:
                    for ci, t in pt_temp_row.items():
                        if ci - 1 < len(row):
                            p = _to_float(row[ci - 1])
                            if p is not None:
                                pc.pt_ratings.append(PTRating(temperature_c=t, max_pressure_barg=p))
                # hydrotest often last column
                last = _to_float(row[-1])
                if last is not None and last > 0:
                    pc.attributes.setdefault("hydrotest_pressure_barg", _attr(last, unit="barg"))
                continue

        # Pipe data: per-size rows (OD, Schedule, WT) + scalar rows (Type, MOC, Ends)
        if section == "pipe_data":
            label = (str(cell_b).strip().lower() if isinstance(cell_b, str) else "")
            if "code" in str(cell_a or "").lower():
                pc.attributes["pipe_std"] = _attr(row[2] if len(row) > 2 else None)
                continue
            if label in ("o.d. (mm)", "od (mm)", "o.d.(mm)") and size_cols:
                _ensure_pipe_rows(pc, size_cols)
                for ci, nps in size_cols.items():
                    _set_pipe_field(pc, nps, "od_mm", _to_float(row[ci - 1]))
                continue
            if "schedule" in label and size_cols:
                _ensure_pipe_rows(pc, size_cols)
                for ci, nps in size_cols.items():
                    _set_pipe_field(pc, nps, "schedule_val", _coerce_str(row[ci - 1]))
                continue
            if "w.t" in label or "wall" in label:
                _ensure_pipe_rows(pc, size_cols)
                for ci, nps in size_cols.items():
                    _set_pipe_field(pc, nps, "wall_thickness_mm", _to_float(row[ci - 1]))
                continue
            if label == "type":
                pc.attributes["pipe_type"] = _attr(row[2] if len(row) > 2 else None)
                continue
            if label == "moc":
                pc.attributes["pipe_moc"] = _attr(row[2] if len(row) > 2 else None)
                continue
            if label == "ends":
                pc.attributes["pipe_ends"] = _attr(row[2] if len(row) > 2 else None)
                continue

        # Flange
        if section == "flange":
            key = _normalize_key(str(cell_a or cell_b or ""))
            value = row[2] if len(row) > 2 else None
            if key and not _is_blank(value):
                pc.attributes[f"flange_{key}"] = _attr(value)
            continue

        # Bolts/Nuts/Gaskets
        if section == "bolts":
            key = _normalize_key(str(cell_a or ""))
            value = row[2] if len(row) > 2 else None
            if key and not _is_blank(value):
                pc.attributes[f"bolting_{key}"] = _attr(value)
            continue

        # Valves
        if section == "valves":
            label_a = str(cell_a or "").strip().lower()
            label_b_l = str(cell_b or "").strip().lower()
            if label_a == "rating":
                pc.attributes.setdefault("valve_rating_label", _attr(row[2] if len(row) > 2 else None))
                continue
            vt = VALVE_TYPE_MAP.get(label_b_l) or VALVE_TYPE_MAP.get(label_a)
            if vt and size_cols:
                # PMS convention: a cell value at size N applies forward until
                # the next non-blank cell (which may add or replace codes).
                ordered = sorted(size_cols.items())  # [(col, nps), ...]
                segments: List[Tuple[float, float, List[str]]] = []
                current_codes: List[str] = []
                seg_start: Optional[float] = None
                for idx, (ci, nps) in enumerate(ordered):
                    cell = row[ci - 1]
                    if not _is_blank(cell):
                        if seg_start is not None and current_codes:
                            prev_nps = ordered[idx - 1][1] if idx > 0 else nps
                            segments.append((seg_start, prev_nps, current_codes))
                        current_codes = [c.strip() for c in re.split(r"[,/]+", str(cell)) if c.strip()]
                        seg_start = nps
                if seg_start is not None and current_codes:
                    segments.append((seg_start, ordered[-1][1], current_codes))
                # merge consecutive segments with identical code sets
                for lo, hi, codes in segments:
                    pc.valve_assignments.append(ValveAssignment(
                        valve_type=vt,
                        nps_min=lo,
                        nps_max=hi,
                        vds_codes=codes,
                        raw_cell_value=", ".join(codes),
                    ))
            continue

        if section == "notes":
            for v in row:
                if not _is_blank(v):
                    notes.append(str(v).strip())
            continue

        # Fallback — stash in extra
        pc.extra.setdefault(section or "misc", []).append([v for v in row if not _is_blank(v)])

    if notes:
        pc.extra["notes"] = notes

    # Promote spec_code from header attributes if present
    sc_attr = pc.attributes.pop("spec_code", None)
    if sc_attr and isinstance(sc_attr.raw, str):
        pc.spec_code = sc_attr.raw

    return pc


def _coerce_str(v: Any) -> Optional[str]:
    if _is_blank(v):
        return None
    return str(v).strip()


def _ensure_pipe_rows(pc: PipingClass, size_cols: Dict[int, float]) -> None:
    existing = {row.nps_inch for row in pc.pipe_schedule}
    for nps in size_cols.values():
        if nps not in existing:
            pc.pipe_schedule.append(PipeScheduleRow(nps_inch=nps))
            existing.add(nps)


def _set_pipe_field(pc: PipingClass, nps: float, field: str, value: Any) -> None:
    for row in pc.pipe_schedule:
        if row.nps_inch == nps:
            setattr(row, field, value)
            return


# ---- entry point ------------------------------------------------------------

def parse_xlsx(path: Path, project_id: str, project_name: Optional[str] = None) -> ProjectPMS:
    wb = openpyxl.load_workbook(path, data_only=True)
    classes: Dict[str, PipingClass] = {}
    for ws in wb.worksheets:
        title = ws.title.strip()
        # Skip non-PMS sheets like branch charts
        if re.search(r"chart", title, re.I):
            continue
        try:
            pc = parse_sheet(ws)
        except Exception:
            continue
        if not pc.attributes and not pc.valve_assignments:
            continue
        classes[pc.spec_code] = pc

    meta = ProjectMetadata(
        project_id=project_id,
        name=project_name or project_id,
        source_file=path.name,
        uploaded_at=datetime.now(timezone.utc).isoformat(),
        status="draft",
    )
    return ProjectPMS(metadata=meta, piping_classes=classes)
